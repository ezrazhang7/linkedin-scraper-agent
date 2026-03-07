"""
json_to_xlsx.py
Usage: python json_to_xlsx.py <input.json> [output.xlsx]
Converts cleaned LinkedIn saved posts JSON → formatted opportunities spreadsheet.
"""

import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# ── Category colours (background) ─────────────────────────────────────────────
CAT_COLORS = {
    "job_opportunity": "D6EAF8",   # blue
    "event":           "D5F5E3",   # green
    "article":         "FEF9E7",   # yellow
    "tool":            "F9EBEA",   # pink
    "person":          "EBE8F9",   # lavender
    "other":           "F2F3F4",   # grey
}

CAT_LABELS = {
    "job_opportunity": "💼 Opportunity",
    "event":           "📅 Event",
    "article":         "📰 Article",
    "tool":            "🛠 Tool",
    "person":          "👤 Person",
    "other":           "📌 Other",
}

HEADER_FILL  = PatternFill("solid", start_color="1B2631")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT    = Font(name="Arial", size=9)
THIN         = Side(style="thin", color="D5D8DC")
BORDER       = Border(bottom=THIN, right=THIN)
WRAP         = Alignment(wrap_text=True, vertical="top")
CENTER       = Alignment(horizontal="center", vertical="top")


def best_link(post: dict) -> str:
    """Return the best clickable external link, or the LinkedIn post URL."""
    ext = post.get("externalLinks") or []
    # prefer non-lnkd.in links
    real = [u for u in ext if "lnkd.in" not in u and "bit.ly" not in u]
    if real:
        return real[0]
    if ext:
        return ext[0]
    return post.get("postUrl", "")


def all_links_str(post: dict) -> str:
    ext = post.get("externalLinks") or []
    return "\n".join(ext) if ext else ""


def build_spreadsheet(data: list, out_path: str):
    wb = Workbook()

    # ── Sheet 1: All posts ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Posts"

    cols = [
        ("#",          5),
        ("Category",   16),
        ("Date",       12),
        ("Summary",    55),
        ("Post Text",  55),
        ("Link",       45),
        ("All External Links", 45),
        ("LinkedIn Post URL", 50),
    ]

    # Header row
    for c, (label, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Data rows
    for row_i, post in enumerate(data, 2):
        cat    = post.get("category", "other")
        color  = CAT_COLORS.get(cat, "F2F3F4")
        fill   = PatternFill("solid", start_color=color)
        label  = CAT_LABELS.get(cat, cat)
        link   = best_link(post)
        date   = post.get("date") or post.get("parsedDate") or ""
        if date and len(date) > 10:
            date = date[:10]

        values = [
            row_i - 1,
            label,
            date,
            post.get("summary", ""),
            (post.get("text") or "")[:500],
            link,
            all_links_str(post),
            post.get("postUrl", ""),
        ]

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=c, value=val)
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = CENTER if c in (1, 3) else WRAP

            # Make link columns clickable
            if c == 6 and val and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="1F618D", underline="single")

        ws.row_dimensions[row_i].height = 60

    # Column widths
    for c, (_, width) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = width

    # ── Sheet 2: Opportunities only ───────────────────────────────────────────
    ws2 = wb.create_sheet("Opportunities")
    opp_cols = [
        ("#",        5),
        ("Date",     12),
        ("Summary",  65),
        ("Link",     50),
        ("Post URL", 50),
    ]

    for c, (label, _) in enumerate(opp_cols, 1):
        cell = ws2.cell(row=1, column=c, value=label)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    ws2.row_dimensions[1].height = 20
    ws2.freeze_panes = "A2"

    opp_fill = PatternFill("solid", start_color=CAT_COLORS["job_opportunity"])
    opps = [p for p in data if p.get("category") == "job_opportunity"]

    for row_i, post in enumerate(opps, 2):
        date = post.get("date") or post.get("parsedDate") or ""
        if date and len(date) > 10:
            date = date[:10]
        link = best_link(post)

        values = [row_i - 1, date, post.get("summary", ""), link, post.get("postUrl", "")]
        for c, val in enumerate(values, 1):
            cell = ws2.cell(row=row_i, column=c, value=val)
            cell.font      = BODY_FONT
            cell.fill      = opp_fill
            cell.border    = BORDER
            cell.alignment = CENTER if c in (1, 2) else WRAP
            if c == 4 and val and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="1F618D", underline="single")

        ws2.row_dimensions[row_i].height = 45

    for c, (_, width) in enumerate(opp_cols, 1):
        ws2.column_dimensions[get_column_letter(c)].width = width

    # ── Sheet 3: Events ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Events")
    ev_fill = PatternFill("solid", start_color=CAT_COLORS["event"])
    events = [p for p in data if p.get("category") == "event"]

    for c, (label, _) in enumerate(opp_cols, 1):
        cell = ws3.cell(row=1, column=c, value=label)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    ws3.row_dimensions[1].height = 20
    ws3.freeze_panes = "A2"

    for row_i, post in enumerate(events, 2):
        date = post.get("date") or post.get("parsedDate") or ""
        if date and len(date) > 10:
            date = date[:10]
        link = best_link(post)
        values = [row_i - 1, date, post.get("summary", ""), link, post.get("postUrl", "")]
        for c, val in enumerate(values, 1):
            cell = ws3.cell(row=row_i, column=c, value=val)
            cell.font      = BODY_FONT
            cell.fill      = ev_fill
            cell.border    = BORDER
            cell.alignment = CENTER if c in (1, 2) else WRAP
            if c == 4 and val and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="1F618D", underline="single")
        ws3.row_dimensions[row_i].height = 45

    for c, (_, width) in enumerate(opp_cols, 1):
        ws3.column_dimensions[get_column_letter(c)].width = width

    wb.save(out_path)
    print(f"✅ Saved → {out_path}")
    print(f"   Total posts : {len(data)}")
    print(f"   Opportunities: {len(opps)}")
    print(f"   Events       : {len(events)}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python json_to_xlsx.py <input.json> [output.xlsx]")
        sys.exit(1)

    json_path = sys.argv[1]
    out_path  = sys.argv[2] if len(sys.argv) > 2 else json_path.replace(".json", ".xlsx")

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    build_spreadsheet(data, out_path)